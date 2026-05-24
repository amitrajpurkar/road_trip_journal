"""
export_to_csv.py
----------------
Reads each sheet from trip_entries.xlsx and writes the data back to the
corresponding CSV file. Run this from the rtrip_2026 folder:

    python OUTPUTS/export_to_csv.py

Requirements: openpyxl  (pip install openpyxl)
"""

import csv, os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

SHEET_TO_CSV = {
    "Itinerary":    os.path.join(BASE, "ITINERARY", "itinerary.csv"),
    "Journal":      os.path.join(BASE, "JOURNAL", "daily-entries", "journal_entries.csv"),
    "Bookings":     os.path.join(BASE, "PLANNING", "bookings", "bookings.csv"),
    "Budget":       os.path.join(BASE, "PLANNING", "budget", "budget.csv"),
    "Packing":      os.path.join(BASE, "PLANNING", "packing", "packing.csv"),
    "Route":        os.path.join(BASE, "PLANNING", "route", "route.csv"),
    "Destinations": os.path.join(BASE, "PLANNING", "destinations", "destinations.csv"),
}

XLSX = os.path.join(BASE, "OUTPUTS", "trip_entries.xlsx")

def export():
    wb = load_workbook(XLSX, data_only=True)
    exported = []

    for sheet_name, csv_path in SHEET_TO_CSV.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [skip] Sheet '{sheet_name}' not found in workbook")
            continue

        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # skip completely empty rows
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                rows.append([str(c) if c is not None else "" for c in row])

        if not rows:
            print(f"  [skip] Sheet '{sheet_name}' is empty")
            continue

        os.makedirs(os.path.dirname(csv_path), exist_ok=True)
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(rows)

        exported.append(sheet_name)
        print(f"  ✓  {sheet_name:14s} → {os.path.relpath(csv_path, BASE)}")

    print(f"\nDone — {len(exported)} CSV files updated.")

if __name__ == "__main__":
    print(f"Exporting from: {os.path.relpath(XLSX, BASE)}\n")
    export()
